from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
from deep_translator import GoogleTranslator
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure
from datetime import datetime, timezone
import logging
import os
import subprocess
import shutil
import speech_recognition as sr
from PIL import Image, ImageOps
import pytesseract
import numpy as np
from io import BytesIO
import base64
try:
    from rapidfuzz import process as rf_process, fuzz as rf_fuzz  # type: ignore
except Exception:
    rf_process = None
    rf_fuzz = None

# Set Tesseract path explicitly
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

from werkzeug.utils import secure_filename
import cv2
app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'wav', 'mp3', 'ogg', 'm4a', 'flac', 'aac', 'webm'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'bmp', 'tif', 'tiff', 'webp'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

CORS(app, resources={
    r"/translate": {"origins": "*"},
    r"/transcribe": {"origins": "*"},
    r"/ocr": {"origins": "*"},
    r"/ocr_receipt": {"origins": "*"},
    r"/extract_fields": {"origins": "*"}
})

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Check ffmpeg availability once at startup
FFMPEG_AVAILABLE = shutil.which('ffmpeg') is not None
if FFMPEG_AVAILABLE:
    logger.info("✅ ffmpeg detected on PATH")
else:
    logger.warning("⚠️ ffmpeg not found on PATH. Only WAV/FLAC inputs will be processed without conversion.")

# Check Tesseract availability once at startup
TESSERACT_AVAILABLE = shutil.which('tesseract') is not None or os.path.exists(pytesseract.pytesseract.tesseract_cmd)
if TESSERACT_AVAILABLE:
    logger.info("✅ Tesseract OCR detected")
else:
    logger.warning("⚠️ Tesseract OCR not found. /ocr endpoint will fail until installed.")

LANGUAGES = {
    "Tamil": "ta",
    "English": "en",
    "Hindi": "hi",
    "French": "fr",
    "German": "de",
    "Japanese": "ja"
}

# ---------------- HELPERS ----------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

def _order_points(pts):
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect

def four_point_transform(image, pts):
    rect = _order_points(pts)
    (tl, tr, br, bl) = rect
    widthA = np.linalg.norm(br - bl)
    widthB = np.linalg.norm(tr - tl)
    maxWidth = int(max(widthA, widthB))
    heightA = np.linalg.norm(tr - br)
    heightB = np.linalg.norm(tl - bl)
    maxHeight = int(max(heightA, heightB))
    dst = np.array([
        [0, 0],
        [maxWidth - 1, 0],
        [maxWidth - 1, maxHeight - 1],
        [0, maxHeight - 1]
    ], dtype="float32")
    M = cv2.getPerspectiveTransform(rect, dst)
    warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))
    return warped

def _read_image_exif_oriented(image_path):
    """Read image with EXIF orientation applied, return OpenCV BGR array."""
    pil = Image.open(image_path)
    try:
        pil = ImageOps.exif_transpose(pil)
    except Exception:
        pass
    return cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2BGR)

def preprocess_receipt(image_path):
    """Return a cleaned image suitable for OCR plus multiple enhanced variants."""
    image = _read_image_exif_oriented(image_path)
    if image is None:
        raise ValueError("Could not read image")

    # Resize for consistent processing
    target_height = 1600
    if image.shape[0] > target_height:
        scale = target_height / image.shape[0]
        image = cv2.resize(image, (int(image.shape[1] * scale), target_height))

    # Perspective correction (same as before)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.bilateralFilter(gray, 11, 17, 17)
    edged = cv2.Canny(blur, 30, 200)
    contours, _ = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    contours = sorted(contours, key=cv2.contourArea, reverse=True)[:10]
    screenCnt = None
    for c in contours:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4:
            screenCnt = approx
            break
    if screenCnt is not None:
        image = four_point_transform(image, screenCnt.reshape(4, 2))

    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Apply noise removal (median blur)
    median = cv2.medianBlur(gray, 3)
    
    # Deskew using minAreaRect
    coords = np.column_stack(np.where(median < 200))
    angle = 0.0
    if coords.size > 0:
        rect = cv2.minAreaRect(coords)
        angle = rect[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
    # Fallback to Tesseract OSD if minAreaRect fails
    if abs(angle) <= 0.5:
        try:
            osd = pytesseract.image_to_osd(median)
            import re as _re
            m = _re.search(r"Rotate: (\d+)", osd)
            if m:
                rdeg = int(m.group(1))
                if rdeg in (90,180,270):
                    angle = -rdeg
        except Exception:
            pass

    # Rotate the receipt upright if needed
    if abs(angle) > 0.5:
        (h, w) = image.shape[:2]
        M = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1.0)
        image = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        # Update grayscale and median blur after rotation
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        median = cv2.medianBlur(gray, 3)

    # Apply thresholding (OTSU) for better text contrast
    otsu = cv2.threshold(median, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    
    # Additional processing for variants
    # Sharpen (unsharp mask)
    gblur = cv2.GaussianBlur(gray, (0, 0), sigmaX=1.0)
    sharp = cv2.addWeighted(gray, 1.5, gblur, -0.5, 0)

    # Adaptive thresholding as another variant
    adaptive = cv2.adaptiveThreshold(median, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 15)

    # Morphology to separate lines (helps OCR grouping)
    kernel_h = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 1))
    sep = cv2.morphologyEx(otsu, cv2.MORPH_OPEN, kernel_h)

    # Small noise removal
    clean = cv2.morphologyEx(otsu, cv2.MORPH_OPEN, np.ones((1, 1), np.uint8))

    # Return primary and variants for strategy OCR
    variants = [clean, adaptive, sharp, otsu, sep]
    return variants[0], variants

GENERAL_CONFIG = "--oem 1 --psm 6"
DIGIT_CONFIG = "--oem 1 --psm 7 -c tessedit_char_whitelist=0123456789.,₹$SAR€"

def ocr_with_strategies(img_variants):
    """Run OCR with multiple PSMs and configs; return best full text and boxes."""
    best_text = ""
    best_len = 0
    all_boxes = []
    best_conf = -1.0
    
    # Ensure we use the preprocessed deskewed image (first variant) with the general config first
    deskewed_img = img_variants[0]
    d = pytesseract.image_to_data(deskewed_img, lang='ara+eng', config=GENERAL_CONFIG, output_type=pytesseract.Output.DICT)
    lines = _reconstruct_lines_from_ocr(d)
    text = _normalize_arabic_digits('\n'.join(lines))
    # average confidence
    confs = [float(c) for c in d['conf'] if c not in ('-1', '')]
    avg_conf = sum(confs)/len(confs) if confs else -1
    best_len = len(text)
    best_text = text
    best_conf = avg_conf
    # capture boxes for this best run
    all_boxes = []
    for i in range(len(d['text'])):
        t = (d['text'][i] or '').strip()
        if not t:
            continue
        x, y, w, h = d['left'][i], d['top'][i], d['width'][i], d['height'][i]
        conf = float(d['conf'][i]) if d['conf'][i] not in ('-1', '') else -1
        all_boxes.append({'x': x, 'y': y, 'w': w, 'h': h, 'text': t, 'conf': conf})
    
    # Try other variants and configurations if needed
    for variant in img_variants[1:]:
        for cfg in (GENERAL_CONFIG, "--oem 1 --psm 4"):
            d = pytesseract.image_to_data(variant, lang='ara+eng', config=cfg, output_type=pytesseract.Output.DICT)
            lines = _reconstruct_lines_from_ocr(d)
            text = _normalize_arabic_digits('\n'.join(lines))
            # average confidence
            confs = [float(c) for c in d['conf'] if c not in ('-1', '')]
            avg_conf = sum(confs)/len(confs) if confs else -1
            if len(text) > best_len or (len(text) >= best_len*0.9 and avg_conf > best_conf):
                best_len = len(text)
                best_text = text
                best_conf = avg_conf
                # capture boxes for this best run
                all_boxes = []
                for i in range(len(d['text'])):
                    t = (d['text'][i] or '').strip()
                    if not t:
                        continue
                    x, y, w, h = d['left'][i], d['top'][i], d['width'][i], d['height'][i]
                    conf = float(d['conf'][i]) if d['conf'][i] not in ('-1', '') else -1
                    all_boxes.append({'x': x, 'y': y, 'w': w, 'h': h, 'text': t, 'conf': conf})

        # Specialized passes
        price_cfg = DIGIT_CONFIG
        _ = pytesseract.image_to_string(variant, lang='eng', config=price_cfg)
        name_cfg = "--psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz \n.:-"
        _ = pytesseract.image_to_string(variant, lang='eng', config=name_cfg)

    # Optional EasyOCR fallback
    try:
        import easyocr  # type: ignore
    except Exception:
        logger.warning("easyocr not installed: OCR fallback disabled")
    else:
        try:
            reader = easyocr.Reader(['en', 'ar'], gpu=False)
            result = reader.readtext(img_variants[0])
            easy_lines = [r[1] for r in result]
            easy_text = _normalize_arabic_digits('\n'.join(easy_lines))
            if len(easy_text) > best_len * 0.7 and len(easy_text) > best_len and best_conf < 50:
                best_text = easy_text
        except Exception:
            pass

    return best_text, all_boxes

def analyze_layout_from_boxes(boxes):
    """Heuristic: determine header/top lines, items, totals by vertical bands.
    Bands are computed from box coordinates themselves to avoid size mismatch.
    """
    if not boxes:
        return [], [], []
    max_y = max((b['y'] + b['h']) for b in boxes if 'y' in b and 'h' in b)
    if max_y <= 0:
        return [], [], []
    top_y = max_y * 0.2
    bottom_y = max_y * 0.8
    header_words = [b['text'] for b in boxes if b['y'] <= top_y]
    total_words = [b['text'] for b in boxes if b['y'] >= bottom_y]
    middle_words = [b['text'] for b in boxes if top_y < b['y'] < bottom_y]
    return header_words, middle_words, total_words

def fuzzy_fix_keyword(word):
    if not rf_process:
        return word
    choices = ["TOTAL","Grand Total","Amount Due","Balance Due","Subtotal","VAT","Tax","Hospital","Clinic"]
    m = rf_process.extractOne(word, choices, scorer=rf_fuzz.WRatio)
    if m and m[1] >= 85:
        return m[0]
    return word

def extract_fields_advanced(full_text, boxes, image_height=None):
    """Combine regex + layout + fuzzy to extract store/date/items/totals."""
    fields = parse_medical_fields(full_text)
    import re
    lines = [ln.strip() for ln in full_text.split('\n') if ln.strip()]
    # Fuzzy correct obvious headings in text
    if rf_process:
        lines = [fuzzy_fix_keyword(ln) for ln in lines]
    # Layout hints
    header_words, middle_words, total_words = analyze_layout_from_boxes(boxes)
    header_line = ' '.join(header_words[:30]).strip()
    if header_line:
        # Try to split header into a probable name and hospital/clinic
        if not fields.get('hospital'):
            fields['hospital'] = header_line[:120]
        # Patient name heuristics near top
        name_patterns = [r"\b(Name|Patient Name)\b\s*[:：-]?\s*(.+)", r"اسم\s*المريض\s*[:：-]?\s*(.+)", r"\bالاسم\b\s*[:：-]?\s*(.+)"]
        for p in name_patterns:
            m = re.search(p, full_text, re.IGNORECASE)
            if m:
                cand = m.group(len(m.groups()))
                fields['patient_name'] = cand.split('\n')[0][:80].strip()
                break
    # Date robust regex
    date_re = re.compile(r"(\b\d{1,2}[\-/]\d{1,2}[\-/]\d{2,4}\b|\b\d{4}[\-/]\d{1,2}[\-/]\d{1,2}\b|\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4}\b)")
    for ln in lines:
        m = date_re.search(ln)
        if m:
            # reject noisy context likely not a date
            if not re.search(r"\b(pcs|piece|register|store|bo\d+)\b", ln, re.IGNORECASE):
                fields['date'] = m.group(0)
            break
    # Items from middle band (exclude totals keywords)
    price_re = re.compile(r"(\d+[.,]\d{2}|SAR\s*\d+|\d+\s*SAR|₹\s*\d+|\$\s*\d+|ريال)\b", re.IGNORECASE)
    item_lines = []
    for ln in middle_words:
        t = str(ln)
        if price_re.search(t) and any(c.isalpha() for c in t):
            if re.search(r"\b(total|amount|balance|vat|tax)\b", t, re.IGNORECASE):
                continue
            item_lines.append(t)
    if item_lines:
        fields['medicines'] = item_lines[:120]
    # Totals from bottom band or text
    totals = []
    totals_keys = ["TOTAL","Grand Total","Amount Due","Balance Due","Subtotal","VAT","Tax"]
    for w in total_words + [' '.join(lines[-6:])]:
        for k in totals_keys:
            if k.lower() in str(w).lower():
                # require amount on the same line
                m = re.search(r"(?:₹|\$|SAR)?\s*\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})\b", str(w))
                if m:
                    totals.append(f"{k}: {m.group(0).strip()}")
    if totals:
        fields['notes'] = ' | '.join(dict.fromkeys(totals))
    
    # Add field_type to boxes for better visualization in frontend
    for box in boxes:
        box_text = box.get('text', '').lower()
        
        # Identify field types based on content
        if fields.get('patient_name') and fields['patient_name'].lower() in box_text:
            box['field_type'] = 'patient_name'
        elif fields.get('date') and fields['date'].lower() in box_text:
            box['field_type'] = 'date'
        elif fields.get('doctor') and fields['doctor'].lower() in box_text:
            box['field_type'] = 'doctor'
        elif fields.get('hospital') and fields['hospital'].lower() in box_text:
            box['field_type'] = 'hospital'
        elif any(med.lower() in box_text for med in fields.get('medicines', []) if med):
            box['field_type'] = 'medicine'
        elif any(k.lower() in box_text for k in totals_keys):
            box['field_type'] = 'total'
    
    return fields

def image_to_base64(img_array):
    pil_img = Image.fromarray(img_array)
    buffer = BytesIO()
    pil_img.save(buffer, format='PNG')
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

def _reconstruct_lines_from_ocr(ocr_dict):
    """Group pytesseract image_to_data output into full text lines in reading order."""
    try:
        n = len(ocr_dict.get('text', []))
        rows = []
        for i in range(n):
            txt = (ocr_dict['text'][i] or '').strip()
            if txt == '':
                continue
            rows.append({
                'block': int(ocr_dict['block_num'][i]),
                'par': int(ocr_dict['par_num'][i]),
                'line': int(ocr_dict['line_num'][i]),
                'word': int(ocr_dict['word_num'][i]),
                'left': int(ocr_dict['left'][i]),
                'top': int(ocr_dict['top'][i]),
                'text': txt
            })
        # sort by block, paragraph, line, then x
        rows.sort(key=lambda r: (r['block'], r['par'], r['line'], r['left']))
        lines = []
        current_key = None
        current_parts = []
        for r in rows:
            key = (r['block'], r['par'], r['line'])
            if key != current_key:
                if current_parts:
                    lines.append(' '.join(current_parts))
                current_key = key
                current_parts = [r['text']]
            else:
                current_parts.append(r['text'])
        if current_parts:
            lines.append(' '.join(current_parts))
        # Deduplicate and keep order
        seen = set()
        unique_lines = []
        for ln in lines:
            if ln in seen:
                continue
            seen.add(ln)
            unique_lines.append(ln)
        return unique_lines
    except Exception:
        # Fallback to joining words
        return [w for w in ocr_dict.get('text', []) if (w or '').strip()]

def _normalize_arabic_digits(s: str) -> str:
    """Convert Arabic-Indic digits to ASCII digits so regex works consistently."""
    trans = str.maketrans(
        '٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹',
        '01234567890123456789'
    )
    return s.translate(trans)

def parse_medical_fields(text):
    import re
    from datetime import datetime
    data = {
        'patient_name': '',
        'date': '',
        'doctor': '',
        'hospital': '',
        'medicines': [],
        'notes': ''
    }

    # Normalize spacing
    normalized = re.sub(r"\t+", " ", text)
    lines = [ln.strip() for ln in normalized.splitlines() if ln.strip()]

    # Helper: normalize date to YYYY-MM-DD if possible
    def _normalize_date(s: str) -> str:
        s = s.strip()
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%y", "%d/%m/%y", "%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
        return s

    # Patient Name → first line with ≥2 words (not doctor/hospital)
    for i, ln in enumerate(lines[:10]):  # Check first 10 lines
        words = ln.split()
        if len(words) >= 2 and not any(keyword in ln.lower() for keyword in ['dr', 'doctor', 'hospital', 'clinic', 'invoice', 'receipt']):
            data['patient_name'] = ln
            break

    # Date → regex: \d{2}[/-]\d{2}[/-]\d{4}
    date_found = ''
    date_pat = re.compile(r"(\b\d{1,2}[\-/]\d{1,2}[\-/]\d{2,4}\b|\b\d{4}[\-/]\d{1,2}[\-/]\d{1,2}\b)")
    for ln in lines:
        m = date_pat.search(ln)
        if m:
            date_found = _normalize_date(m.group(0))
            break
    if date_found:
        data['date'] = date_found

    # Doctor → any line containing "Dr" or "Doctor"
    for ln in lines:
        if re.search(r"\b(Dr|Doctor|Dr\.|دكتور)\b", ln, re.IGNORECASE):
            data['doctor'] = ln.strip()
            break

    # Hospital/Clinic → fuzzy match "hospital" or "clinic"
    if rf_process:  # Check if rapidfuzz is available
        hospital_found = False
        for ln in lines:
            # Use rapidfuzz for fuzzy matching
            match = rf_process.extractOne(ln.lower(), ["hospital", "clinic", "medical center", "مستشفى", "عيادة"], 
                                        scorer=rf_fuzz.WRatio)
            if match and match[1] >= 70:  # 70% similarity threshold
                data['hospital'] = ln.strip()
                hospital_found = True
                break
        
        # If no match found, try to infer from header lines
        if not hospital_found:
            header_lines = lines[:5]
            for ln in header_lines:
                if len(ln.split()) <= 5 and not re.search(r"\b(invoice|receipt|bill|patient|doctor)\b", ln.lower()):
                    data['hospital'] = ln.strip()
                    break
    else:
        # Fallback if rapidfuzz is not available
        m = re.search(r"\b(Hospital|Clinic|Medical Center|مستشفى|عيادة)[^\n]*", normalized, re.IGNORECASE)
        if m:
            data['hospital'] = m.group(0).strip()

    # Medicines → lines that are not totals/amounts/bills
    meds = []
    # Skip lines that look like totals or headers
    skip_patterns = [r"\b(total|subtotal|amount|balance|vat|tax|invoice|receipt)\b", 
                    r"\b(الإجمالي|المجموع|ضريبة|فاتورة)\b"]
    
    # Find potential medicine lines
    for ln in lines:
        if len(ln.split()) >= 2:  # At least 2 words
            skip = False
            for pattern in skip_patterns:
                if re.search(pattern, ln, re.IGNORECASE):
                    skip = True
                    break
            if not skip:
                # Check if it looks like a medicine entry
                if re.search(r"[A-Za-z]{3,}", ln) and not ln == data['patient_name'] and not ln == data['doctor'] and not ln == data['hospital']:
                    meds.append(ln)
    
    if meds:
        data['medicines'] = meds[:100]

    # Notes → leftover lines after the above
    leftover_notes = []
    for ln in lines:
        if (ln != data['patient_name'] and 
            ln != data['doctor'] and 
            ln != data['hospital'] and 
            ln not in data['medicines'] and
            ln != data['date']):
            # Check if it's a total or important information
            if re.search(r"\b(total|amount|balance|due|vat|tax)\b", ln, re.IGNORECASE):
                leftover_notes.append(ln)
    
    if leftover_notes:
        data['notes'] = "\n".join(leftover_notes[:10])  # Limit to 10 lines

    return data

def convert_to_wav(input_path, output_path):
    """Convert any audio file to WAV using ffmpeg"""
    try:
        if not FFMPEG_AVAILABLE:
            logger.error("FFmpeg unavailable: cannot convert %s to wav", input_path)
            return False
        result = subprocess.run([
            'ffmpeg', '-y', '-i', input_path,
            '-ar', '16000', '-ac', '1', '-c:a', 'pcm_s16le', output_path
        ], capture_output=True, text=True)
        if result.returncode != 0:
            logger.error(f"FFmpeg error: {result.stderr}")
            return False
        if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
            logger.error("FFmpeg failed: output file missing or empty")
            return False
        logger.info(f"Conversion successful: {output_path}")
        return True
    except Exception as e:
        logger.error(f"convert_to_wav exception: {e}")
        return False

# ---------------- MONGO ----------------
MONGODB_AVAILABLE = False
try:
    client = MongoClient('mongodb://localhost:27017/', serverSelectionTimeoutMS=5000)
    client.admin.command('ping')
    db = client['translation_db']
    MONGODB_AVAILABLE = True
    logger.info("✅ Connected to MongoDB")
except ConnectionFailure as e:
    logger.warning(f"⚠️ MongoDB not available: {e}")

# ---------------- ROUTES ----------------
@app.route('/')
def home():
    return render_template('index.html', languages=LANGUAGES)

@app.route('/translate', methods=['POST'])
def translate():
    data = request.get_json()
    if not data or 'text' not in data or 'language' not in data:
        return jsonify({'error': 'Missing required fields'}), 400
    
    text = data['text'].strip()
    target_lang = data['language']
    if not text:
        return jsonify({'error': 'No text provided'}), 400
    if target_lang not in LANGUAGES.values():
        return jsonify({'error': 'Unsupported target language'}), 400
    
    try:
        is_arabic = any('\u0600' <= c <= '\u06FF' for c in text)
        translated_text = ""
        if is_arabic and target_lang == 'en':
            sentences = [s.strip() for s in text.split('.') if s.strip()]
            translated_texts = []
            for sentence in sentences:
                try:
                    translator = GoogleTranslator(source='ar', target='en')
                    t = translator.translate(sentence)
                    translated_texts.append(t)
                except:
                    translated_texts.append(sentence)
            translated_text = '. '.join(translated_texts)
        else:
            translator = GoogleTranslator(source='auto', target=target_lang)
            translated_text = translator.translate(text)
        
        if MONGODB_AVAILABLE:
            try:
                db.translations.insert_one({
                    'original_text': text,
                    'translated_text': translated_text,
                    'source_lang': 'ar' if is_arabic else 'auto',
                    'target_lang': target_lang,
                    'timestamp': datetime.now(timezone.utc)
                })
            except Exception as e:
                logger.error(f"Mongo insert failed: {e}")
        
        return jsonify({'original_text': text, 'translated_text': translated_text, 'status': 'success'})
    
    except Exception as e:
        logger.error(f"Translation error: {e}")
        return jsonify({'error': 'Translation failed', 'details': str(e)}), 500

@app.route('/transcribe', methods=['POST'])
def transcribe_audio():
    # Accept common field names or raw body
    audio_file = request.files.get('audio') or request.files.get('file')
    if audio_file is None and request.data:
        # Create a file-like object from raw bytes
        from io import BytesIO
        audio_file = type('RawFile', (), {
            'filename': 'upload.webm',
            'save': lambda self, dst: open(dst, 'wb').write(request.data)
        })()
    if audio_file is None:
        return jsonify({'error': 'No audio provided', 'status': 'error'}), 400
    
    if audio_file.filename == '':
        return jsonify({'error': 'No selected file', 'status': 'error'}), 400
    
    if not allowed_file(audio_file.filename):
        return jsonify({'error': 'Invalid file type', 'details': f"Got '{audio_file.filename}'. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}", 'status': 'error'}), 400
    
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_audio')
    os.makedirs(temp_dir, exist_ok=True)
    
    temp_path = os.path.join(temp_dir, secure_filename(audio_file.filename))
    
    try:
        audio_file.save(temp_path)
        recognizer = sr.Recognizer()
        
        # Convert to WAV if needed
        if not temp_path.lower().endswith('.wav'):
            wav_path = os.path.splitext(temp_path)[0] + '.wav'
            if convert_to_wav(temp_path, wav_path):
                os.remove(temp_path)
                temp_path = wav_path
            else:
                return jsonify({
                    'error': 'Audio conversion failed',
                    'details': 'If uploading webm/mp3/ogg/m4a, ensure ffmpeg is installed and on PATH. Otherwise, upload WAV or FLAC.',
                    'status': 'error'
                }), 400
        
        with sr.AudioFile(temp_path) as source:
            audio_data = recognizer.record(source)
            try:
                text = recognizer.recognize_google(audio_data, language='ar-AR')
            except sr.UnknownValueError:
                try:
                    text = recognizer.recognize_google(audio_data, language='en-US')
                except sr.UnknownValueError:
                    return jsonify({'error': 'Could not understand audio', 'status':'error'}), 400
        
        return jsonify({'text': text, 'status': 'success'})
    
    except Exception as e:
        logger.error(f"Audio processing failed: {e}")
        return jsonify({'error': 'Failed to process audio', 'details': str(e), 'status':'error'}), 500
    
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                logger.error(f"Failed to delete temp file: {e}")

@app.route('/ocr', methods=['POST'])
def ocr_receipt():
    # Accept 'image' or common alternates
    image_file = request.files.get('image') or request.files.get('file') or request.files.get('photo')
    if image_file is None:
        return jsonify({'error': 'No image provided', 'status': 'error'}), 400

    if not TESSERACT_AVAILABLE:
        return jsonify({'error': 'Tesseract OCR is not installed', 'details': 'Install Tesseract and ensure it is on PATH. On Windows: winget install UB-Mannheim.Tesseract-OCR', 'status': 'error'}), 500

    if image_file.filename == '':
        return jsonify({'error': 'No selected image', 'status': 'error'}), 400

    if not allowed_image_file(image_file.filename):
        return jsonify({'error': 'Invalid image type', 'details': f"Got '{image_file.filename}'. Allowed: {', '.join(sorted(ALLOWED_IMAGE_EXTENSIONS))}", 'status': 'error'}), 400

    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_images')
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, secure_filename(image_file.filename))

    try:
        image_file.save(temp_path)
        # Basic preprocessing: convert to grayscale
        img = Image.open(temp_path)
        img = img.convert('L')  # grayscale
        # Optional: further preprocessing could be added here
        lang = request.form.get('lang', 'eng')
        psm = request.form.get('psm')  # optional page segmentation mode
        config = ''
        if psm and str(psm).isdigit():
            config = f'--psm {psm}'
        text = pytesseract.image_to_string(img, lang=lang, config=config)
        text = text.strip()
        return jsonify({'text': text, 'status': 'success'})
    except Exception as e:
        logger.error(f"OCR processing failed: {e}")
        return jsonify({'error': 'Failed to process image', 'details': str(e), 'status': 'error'}), 500
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                logger.error(f"Failed to delete temp image: {e}")

@app.route('/ocr_receipt', methods=['POST'])
def ocr_receipt_advanced():
    if not TESSERACT_AVAILABLE:
        return jsonify({'error': 'Tesseract OCR is not installed', 'status': 'error'}), 500
    f = request.files.get('image') or request.files.get('file') or request.files.get('photo')
    if f is None:
        return jsonify({'error': 'No file provided', 'status': 'error'}), 400
    filename = secure_filename(f.filename or 'upload.png')
    ext = os.path.splitext(filename)[1].lower().lstrip('.')
    if ext == 'pdf':
        # Convert first page of PDF to image
        try:
            from pdf2image import convert_from_bytes
            pages = convert_from_bytes(f.read(), fmt='png')
            if not pages:
                return jsonify({'error': 'Empty PDF', 'status': 'error'}), 400
            buf = BytesIO()
            pages[0].save(buf, format='PNG')
            img_array = np.frombuffer(buf.getvalue(), dtype=np.uint8)
            img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
            temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_images')
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, 'page1.png')
            cv2.imwrite(temp_path, img)
        except Exception as e:
            return jsonify({'error': 'PDF processing failed', 'details': str(e), 'status': 'error'}), 500
    else:
        if not allowed_image_file(filename):
            return jsonify({'error': 'Invalid image type', 'details': f"Allowed: {', '.join(sorted(ALLOWED_IMAGE_EXTENSIONS))} or PDF", 'status': 'error'}), 400
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_images')
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, filename)
        f.save(temp_path)
    try:
        clean, variants = preprocess_receipt(temp_path)
        full_text, boxes = ocr_with_strategies(variants)
        h = variants[0].shape[0]
        fields = extract_fields_advanced(full_text, boxes, h)
        preview_b64 = image_to_base64(variants[0])
        return jsonify({'status': 'success', 'preview_png_base64': preview_b64, 'boxes': boxes, 'fields': fields, 'raw_text': full_text})
    except Exception as e:
        logger.error(f"Advanced OCR failed: {e}")
        return jsonify({'error': 'Advanced OCR failed', 'details': str(e), 'status': 'error'}), 500
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass

@app.route('/export_report', methods=['POST'])
def export_report():
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({'error': 'No JSON provided'}), 400
        fields = payload.get('fields') or {}
        translations = payload.get('translations') or {}
        export_format = (payload.get('format') or 'pdf').lower()

        if export_format == 'pdf':
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            buf = BytesIO()
            c = canvas.Canvas(buf, pagesize=A4)
            width, height = A4
            y = height - 20*mm
            c.setFont('Helvetica-Bold', 16)
            c.drawString(20*mm, y, 'Medical Receipt Report')
            y -= 12*mm
            c.setFont('Helvetica', 11)
            def draw_kv(k, v):
                nonlocal y
                c.drawString(20*mm, y, f"{k}: {v or ''}")
                y -= 8*mm
            draw_kv('Patient Name', fields.get('patient_name'))
            draw_kv('Date', fields.get('date'))
            draw_kv('Doctor', fields.get('doctor'))
            draw_kv('Hospital', fields.get('hospital'))
            meds = '\n'.join(fields.get('medicines') or [])
            draw_kv('Medicines', '')
            for line in meds.split('\n'):
                c.drawString(26*mm, y, line)
                y -= 7*mm
            draw_kv('Notes', fields.get('notes'))
            y -= 6*mm
            c.setFont('Helvetica-Bold', 13)
            c.drawString(20*mm, y, 'Translation (EN)')
            y -= 9*mm
            c.setFont('Helvetica', 11)
            for k, v in translations.items():
                if isinstance(v, list):
                    c.drawString(20*mm, y, f"{k}:")
                    y -= 7*mm
                    for item in v:
                        c.drawString(26*mm, y, f"- {item}")
                        y -= 7*mm
                else:
                    c.drawString(20*mm, y, f"{k}: {v}")
                    y -= 7*mm
                if y < 20*mm:
                    c.showPage(); y = height - 20*mm; c.setFont('Helvetica', 11)
            c.showPage(); c.save()
            buf.seek(0)
            return app.response_class(buf.read(), mimetype='application/pdf', headers={
                'Content-Disposition': 'attachment; filename="receipt_report.pdf"'
            })

        elif export_format in ('xlsx', 'excel'):
            import openpyxl
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Receipt'
            ws.append(['Field', 'Value'])
            ws.append(['Patient Name', fields.get('patient_name', '')])
            ws.append(['Date', fields.get('date', '')])
            ws.append(['Doctor', fields.get('doctor', '')])
            ws.append(['Hospital', fields.get('hospital', '')])
            ws.append(['Medicines', '\n'.join(fields.get('medicines') or [])])
            ws.append(['Notes', fields.get('notes', '')])
            ws2 = wb.create_sheet('Translation')
            ws2.append(['Field', 'English'])
            for k, v in translations.items():
                if isinstance(v, list):
                    ws2.append([k, '\n'.join(v)])
                else:
                    ws2.append([k, v])
            for wsx in (ws, ws2):
                for col in range(1, 3):
                    wsx.column_dimensions[get_column_letter(col)].width = 40
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return app.response_class(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
                'Content-Disposition': 'attachment; filename="receipt_report.xlsx"'
            })
        else:
            return jsonify({'error': 'Unsupported export format'}), 400
    except Exception as e:
        logger.error(f"Export failed: {e}")
        return jsonify({'error': 'Export failed', 'details': str(e)}), 500

@app.route('/extract_fields', methods=['POST'])
def extract_fields_endpoint():
    if not TESSERACT_AVAILABLE:
        return jsonify({'error': 'Tesseract OCR is not installed', 'status': 'error'}), 500
    
    f = request.files.get('image') or request.files.get('file') or request.files.get('photo')
    if f is None:
        return jsonify({'error': 'No file provided', 'status': 'error'}), 400
    
    filename = secure_filename(f.filename or 'upload.png')
    ext = os.path.splitext(filename)[1].lower().lstrip('.')
    
    if not allowed_image_file(filename):
        return jsonify({'error': 'Invalid image type', 'details': f"Allowed: {', '.join(sorted(ALLOWED_IMAGE_EXTENSIONS))}", 'status': 'error'}), 400
    
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_images')
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, filename)
    
    try:
        f.save(temp_path)
        
        # Import and use the extract_fields function from extracter.py
        from extracter import extract_fields
        fields = extract_fields(temp_path)
        
        # Convert image to base64 for preview
        preview_img = cv2.imread(temp_path)
        if preview_img is not None:
            preview_b64 = image_to_base64(preview_img)
        else:
            preview_b64 = None
        
        return jsonify({
            'status': 'success', 
            'fields': fields,
            'preview_png_base64': preview_b64
        })
        
    except Exception as e:
        logger.error(f"Field extraction failed: {e}")
        return jsonify({'error': 'Field extraction failed', 'details': str(e), 'status': 'error'}), 500
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass

if __name__ == '__main__':
    import socket
    # Try to find an available port starting from 5000
    def find_available_port(start_port=5000):
        for port in range(start_port, start_port + 10):
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.bind(('', port))
                    return port
            except OSError:
                continue
        return start_port  # fallback to original port

    port = find_available_port()
    print(f"Starting server on port {port}")

    app.run(
        host='127.0.0.1',
        port=port,
        debug=True,
        use_reloader=False,  # Disable reloader to avoid socket conflicts
        threaded=True
    )
